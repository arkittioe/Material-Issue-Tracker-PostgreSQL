# file: data/report_service.py

import os
import logging
import pandas as pd
from datetime import datetime
from typing import Any, Dict, List, Tuple, Optional, Callable
from sqlalchemy import func, desc

from data.db_session import DBSessionManager
from models import MIVRecord, MTOItem, MTOProgress, SpoolConsumption


class ReportService:
    """
    سرویس تولید گزارشات، تحلیل‌ها و خروجی‌گیری برای پروژه و خطوط.
    بر اساس method_map و منطق اصلی data_manager.py
    """

    def __init__(
        self,
        project_service,                   # باید متدهای get_enriched_line_progress و get_project_line_status_list را داشته باشد
        activity_logger: Callable[..., None],
        session_getter: Callable[[], Any] = DBSessionManager.get_session
    ):
        self.project_service = project_service
        self.log_activity = activity_logger
        self._session_getter = session_getter

    # ===================================================
    # ۱. گزارش کامل یک خط
    # ===================================================
    def get_detailed_line_report(self, project_id: int, line_no: str) -> Dict[str, List]:
        session = self._session_getter()
        try:
            bom = self.project_service.get_enriched_line_progress(
                project_id, line_no, readonly=False
            )

            miv_history_query = session.query(MIVRecord).filter(
                MIVRecord.project_id == project_id,
                MIVRecord.line_no == line_no
            ).order_by(desc(MIVRecord.last_updated)).all()

            miv_history = [
                {
                    "MIV Tag": r.miv_tag,
                    "Registered By": r.registered_by,
                    "Date": r.last_updated.strftime('%Y-%m-%d %H:%M') if r.last_updated else "",
                    "Status": r.status,
                    "Comment": r.comment
                }
                for r in miv_history_query
            ]

            return {"bill_of_materials": bom or [], "miv_history": miv_history}
        except Exception as e:
            logging.error(f"Error in get_detailed_line_report: {e}")
            return {"bill_of_materials": [], "miv_history": []}
        finally:
            session.close()

    # ===================================================
    # ۲. گزارش کسری متریال
    # ===================================================
    def get_shortage_report(self, project_id: int, line_no: Optional[str] = None) -> Dict[str, Any]:
        session = self._session_getter()
        try:
            query = session.query(
                MTOProgress.item_code,
                MTOProgress.description,
                MTOProgress.unit,
                func.sum(MTOProgress.total_qty).label("total_required"),
                func.sum(MTOProgress.used_qty).label("total_used")
            ).filter(MTOProgress.project_id == project_id)

            if line_no:
                query = query.filter(MTOProgress.line_no == line_no)

            results = query.group_by(
                MTOProgress.item_code, MTOProgress.description, MTOProgress.unit
            ).having(
                func.sum(MTOProgress.total_qty) > func.sum(MTOProgress.used_qty)
            ).all()

            data = []
            for row in results:
                remaining = (row.total_required or 0) - (row.total_used or 0)
                progress = (row.total_used / row.total_required * 100) if row.total_required else 0
                data.append({
                    "Item Code": row.item_code or "N/A",
                    "Description": row.description,
                    "Unit": row.unit,
                    "Total Required": round(row.total_required or 0, 2),
                    "Total Used": round(row.total_used or 0, 2),
                    "Remaining": round(remaining, 2),
                    "Progress (%)": round(progress, 2)
                })
            return {"data": data}
        except Exception as e:
            logging.error(f"Error in get_shortage_report: {e}")
            return {"data": []}
        finally:
            session.close()

    # ===================================================
    # ۳. گزارشات تحلیلی برای نمودار
    # ===================================================
    def get_report_analytics(self, project_id: int, report_name: str, **params) -> Dict[str, Any]:
        session = self._session_getter()
        try:
            # توزیع پیشرفت خطوط
            if report_name == 'line_progress_distribution':
                lines = self.project_service.get_project_line_status_list(project_id)
                bins = {"0-25%": 0, "25-50%": 0, "50-75%": 0, "75-99%": 0, "100%": 0}
                for line in lines:
                    p = line.get('Progress (%)', 0)
                    if p < 25:
                        bins["0-25%"] += 1
                    elif p < 50:
                        bins["25-50%"] += 1
                    elif p < 75:
                        bins["50-75%"] += 1
                    elif p < 100:
                        bins["75-99%"] += 1
                    else:
                        bins["100%"] += 1
                return {
                    "title": "توزیع پیشرفت خطوط",
                    "type": "bar",
                    "data": {
                        "labels": list(bins.keys()),
                        "datasets": [{"label": "تعداد خطوط", "data": list(bins.values())}]
                    }
                }

            # مصرف متریال بر اساس نوع
            elif report_name == 'material_usage_by_type':
                results = session.query(
                    MTOItem.item_type,
                    func.sum(MTOProgress.used_qty).label('total_used')
                ).join(
                    MTOProgress, MTOItem.id == MTOProgress.mto_item_id
                ).filter(
                    MTOProgress.project_id == project_id,
                    MTOItem.item_type.isnot(None)
                ).group_by(
                    MTOItem.item_type
                ).order_by(
                    desc('total_used')
                ).limit(10).all()

                return {
                    "title": "۱۰ نوع متریال پرمصرف",
                    "type": "pie",
                    "data": {
                        "labels": [r.item_type for r in results],
                        "datasets": [{"data": [round(r.total_used or 0, 2) for r in results]}]
                    }
                }

            # مصرف اسپول در طول زمان
            elif report_name == 'consumption_over_time':
                results = session.query(
                    func.strftime('%Y-%m-%d', SpoolConsumption.timestamp).label('date'),
                    func.count(SpoolConsumption.id).label('consumption_count')
                ).group_by(
                    'date'
                ).order_by(
                    'date'
                ).all()

                return {
                    "title": "مصرف اسپول در طول زمان",
                    "type": "line",
                    "data": {
                        "labels": [r.date for r in results],
                        "datasets": [{"label": "تعداد مصرف", "data": [r.consumption_count for r in results]}]
                    }
                }

            return {"error": "Report name not found"}
        except Exception as e:
            logging.error(f"Error in get_report_analytics: {e}")
            return {"error": str(e)}
        finally:
            session.close()

    # ===================================================
    # ۴. خروجی گرفتن از داده‌ها (عمومی)
    # ===================================================
    def export_data_to_file(self, data: List[Dict[str, Any]], file_path: str, report_title: str) -> Tuple[bool, str]:
        if not data:
            return False, "داده‌ای برای خروجی گرفتن وجود ندارد."
        file_ext = os.path.splitext(file_path)[1].lower()
        df = pd.DataFrame(data)
        try:
            if file_ext == '.xlsx':
                return self._export_to_excel(df, file_path, report_title)
            elif file_ext == '.pdf':
                from .report_exporter import pdf_export
                return pdf_export(df, file_path, report_title)
            else:
                return False, f"پسوند '{file_ext}' پشتیبانی نمی‌شود."
        except Exception as e:
            logging.error(f"Export error: {e}")
            return False, f"خطا در خروجی: {e}"

    # ===================================================
    # ۵. خروجی گرفتن رکوردهای MIV
    # ===================================================
    def export_miv_records_to_file(self, project_id: int, file_path: str) -> Tuple[bool, str]:
        session = self._session_getter()
        try:
            records = session.query(MIVRecord).filter(
                MIVRecord.project_id == project_id
            ).order_by(desc(MIVRecord.last_updated)).all()

            data = [
                {
                    "MIV Tag": r.miv_tag,
                    "Line No": r.line_no,
                    "Registered By": r.registered_by,
                    "Date": r.last_updated.strftime('%Y-%m-%d %H:%M') if r.last_updated else "",
                    "Status": r.status,
                    "Comment": r.comment
                }
                for r in records
            ]
            return self.export_data_to_file(data, file_path, f"MIV Records - Project {project_id}")
        except Exception as e:
            logging.error(f"Error in export_miv_records_to_file: {e}")
            return False, f"خطا: {e}"
        finally:
            session.close()

    # ===================================================
    # ۶. خروجی گرفتن گزارش کامل یک خط
    # ===================================================
    def export_detailed_line_report_to_file(self, project_id: int, line_no: str, file_path: str) -> Tuple[bool, str]:
        report = self.get_detailed_line_report(project_id, line_no)
        # ادغام BOM و MIV History در دو شیت یا یک شیت
        data = []
        data.append({"Section": "Bill of Materials"})
        data.extend(report.get("bill_of_materials", []))
        data.append({})
        data.append({"Section": "MIV History"})
        data.extend(report.get("miv_history", []))
        return self.export_data_to_file(data, file_path, f"Line {line_no} Report")

    # ===================================================
    # ۷. متد کمکی خروجی اکسل
    # ===================================================
    def _export_to_excel(self, df: pd.DataFrame, file_path: str, report_title: str) -> Tuple[bool, str]:
        with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
            sheet_name = report_title[:30].replace('/', '-')
            df.to_excel(writer, sheet_name=sheet_name, index=False)
            worksheet = writer.sheets[sheet_name]

            from openpyxl.styles import Font, PatternFill
            header_font = Font(bold=True, color='FFFFFF')
            header_fill = PatternFill(start_color='4F81BD', end_color='4F81BD', fill_type='solid')

            for col_num, _ in enumerate(df.columns.values, 1):
                cell = worksheet.cell(row=1, column=col_num)
                cell.font = header_font
                cell.fill = header_fill

            for column_cells in worksheet.columns:
                max_length = max(len(str(cell.value or "")) for cell in column_cells)
                worksheet.column_dimensions[column_cells[0].column_letter].width = max_length + 2

        return True, f"Excel report saved to:\n{file_path}"
